import sqlite3
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
import os

# --- 1. Creating Data Bases---
def create_mock_dbs():
    conn1 = sqlite3.connect("db1.sqlite")
    conn1.execute("DROP TABLE IF EXISTS clients")
    conn1.execute("""
        CREATE TABLE clients (id INTEGER, client_name TEXT, date_exp TEXT, amount REAL, currency TEXT)
    """)
    conn1.executemany("INSERT INTO clients VALUES (?, ?, ?, ?, ?)", [
        (1, "Jan Kowalski", "2026-01-10", 1000, "PLN"),
        (2, "Anna Nowak", "2026-01-15", 500, "PLN"),
        (3, "Mia New", "2026-01-27", 700, "USD"),
        (4, "Tom Smith", "2026-02-05", 300, "USD"),
        (5, "Extra Person", "2026-03-01", 400, "EUR"),
    ])
    conn1.commit()
    conn1.close()

    conn2 = sqlite3.connect("db2.sqlite")
    conn2.execute("DROP TABLE IF EXISTS customers")
    conn2.execute("""
        CREATE TABLE customers (customer_id INTEGER, full_name TEXT, ex_date TEXT, total REAL, currency TEXT)
    """)
    conn2.executemany("INSERT INTO customers VALUES (?, ?, ?, ?, ?)", [
        (101, "Jan Kowalski", "2026-01-10", 1000, "PLN"),
        (102, "Anna Nowak", "2026-01-15", 500, "PLN"),
        (103, "Mia New", "2026-01-27", 700, "EUR"),
        (104, "Tom Smith", "2026-02-05", 300, "USD"),
        (105, "Extra Person", "2026-03-10", 600, "USD"),
    ])
    conn2.commit()
    conn2.close()

    conn3 = sqlite3.connect("db3.sqlite")
    conn3.execute("DROP TABLE IF EXISTS people")
    conn3.execute("""
        CREATE TABLE people (cid TEXT, name TEXT, expiry TEXT, amt REAL, currency TEXT)
    """)
    conn3.executemany("INSERT INTO people VALUES (?, ?, ?, ?, ?)", [
        ("a1", "Jan Kowalski", "2026-01-10", 1000, "EUR"),
        ("a2", "Anna Nowak", "2026-01-15", 500, "PLN"),
        ("a3", "Mia New", "2026-01-27", 700, "USD"),
        ("a4", "Tom Smith", "2026-02-05", 300, "USD"),
        ("a5", "Extra Person", "2026-04-01", 800, "GBP"),
    ])
    conn3.commit()
    conn3.close()

# --- 2. Taking the Data and Preparing DF ---
def fetch_and_compare():
    # DB1
    conn1 = sqlite3.connect("db1.sqlite")
    df1 = pd.read_sql("SELECT * FROM clients", conn1)
    df1 = df1.rename(columns={
        "id": "db1_id",
        "client_name": "db1_full_name",
        "date_exp": "db1_ex_date",
        "amount": "db1_amount",
        "currency": "db1_currency"
    })
    conn1.close()

    # DB2
    conn2 = sqlite3.connect("db2.sqlite")
    df2 = pd.read_sql("SELECT * FROM customers", conn2)
    df2 = df2.rename(columns={
        "customer_id": "db2_id",
        "full_name": "db2_full_name",
        "ex_date": "db2_ex_date",
        "total": "db2_amount",
        "currency": "db2_currency"
    })
    conn2.close()

    # DB3
    conn3 = sqlite3.connect("db3.sqlite")
    df3 = pd.read_sql("SELECT * FROM people", conn3)
    df3 = df3.rename(columns={
        "cid": "db3_id",
        "name": "db3_full_name",
        "expiry": "db3_ex_date",
        "amt": "db3_amount",
        "currency": "db3_currency"
    })
    conn3.close()

    df_all = pd.concat([df1.reset_index(drop=True),
                        df2.reset_index(drop=True),
                        df3.reset_index(drop=True)], axis=1)

    # Comapre Currency
    df_all['comment'] = df_all.apply(
        lambda row: "ok" if row['db1_currency'] == row['db2_currency'] == row['db3_currency'] else "check",
        axis=1
    )

    # Highlight if ex_date < 14 days from today
    today = datetime.today()
    for col in ['db1_ex_date', 'db2_ex_date', 'db3_ex_date']:
        df_all[col] = pd.to_datetime(df_all[col], errors='coerce')
    df_all['highlight'] = df_all[['db1_ex_date','db2_ex_date','db3_ex_date']].apply(
        lambda row: any((today <= d <= today + timedelta(days=14)) for d in row), axis=1
    )

    for col in ['db1_ex_date', 'db2_ex_date', 'db3_ex_date']:
        df_all[col] = df_all[col].dt.strftime('%Y-%m-%d')

    df_all['user_comment'] = ""

    return df_all

# --- 3. Save to Excel ---
def save_to_excel(df):
    today = datetime.today().strftime("%Y%m%d")
    filename = f"report_{today}.xlsx"

    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_font = Font(color="FF0000", bold=True)

    df_to_save = df.drop(columns=['highlight'], errors='ignore')
    df_to_save.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    comment_col_idx = df_to_save.columns.get_loc("comment") + 1
    user_comment_col_idx = df_to_save.columns.get_loc("user_comment") + 1

    for i, row in enumerate(ws.iter_rows(min_row=2), start=0):
        if df.iloc[i]['highlight']:
            for cell in row:
                cell.fill = highlight_fill

        comment_cell = ws.cell(row=i + 2, column=comment_col_idx)
        if comment_cell.value == "check":
            comment_cell.font = red_font

    dv = DataValidation(
        type="list",
        formula1='"clear,contact client,follow up,pending"',
        allow_blank=True
    )
    ws.add_data_validation(dv)
    dv.add(
        f"{ws.cell(row=2, column=user_comment_col_idx).coordinate}:"
        f"{ws.cell(row=ws.max_row, column=user_comment_col_idx).coordinate}"
    )

    wb.save(filename)

def run_pipeline():
    env = os.getenv("ENV", "local")

    print(f"Running pipeline in ENV={env}")

    if env == "local":
        create_mock_dbs()

    df = fetch_and_compare()
    save_to_excel(df)

if __name__ == "__main__":
    run_pipeline()
    print("Excel report generated: report.xlsx")